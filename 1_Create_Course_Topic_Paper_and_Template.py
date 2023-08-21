import openpyxl
import tkinter as tk
from tkinter import filedialog

def edit_excel_file(filename):
    # Load the Excel file
    wb = openpyxl.load_workbook(filename)
    
    # Get the "4. Course Question" sheet
    sheet = wb['4. Course Question']

    # Get the max row number
    max_row = sheet.max_row
    
    # Insert headers in cells L1 and M1
    sheet['L1'] = 'Course Topic Name'
    sheet['M1'] = 'Paper Name'
    
    # Iterate through rows in column K and update formulas in columns L and M
    for row in range(2, max_row + 1):
        if sheet[f'K{row}'].value:
            sheet[f'L{row}'] = f'=VLOOKUP(D{row}, \'1. Course Topic\'!A:C, 3, FALSE)'
            sheet[f'M{row}'] = f'=VLOOKUP(J{row}, \'2. Course Paper\'!A:D, 4, FALSE)'
    
    # Create a new sheet 'Check'
    check_sheet = wb.create_sheet('Check')

    # Load the external Excel file 'Mapping A to I.xlsx'
    external_wb = openpyxl.load_workbook('Mapping A to I.xlsx')
    
    # Get the 'Mapping ID' and 'Cohort' sheets from the external file
    mapping_id_sheet = external_wb['Mapping ID']
    cohort_sheet = external_wb['Cohort']
    
    # Create a new sheet 'Mappings'
    mappings_sheet = wb.create_sheet('Mappings')

    # Copy values from 'Mapping ID' to 'Mappings'
    for row in mapping_id_sheet.iter_rows(min_row=1, values_only=True):
        mappings_sheet.append(row)
    
    # Create a new sheet 'Cohort Mapping'
    cohort_mapping_sheet = wb.create_sheet('Cohort Mapping')
    
    # Copy values from 'Cohort' to 'Cohort Mapping'
    for row in cohort_sheet.iter_rows(min_row=1, values_only=True):
        cohort_mapping_sheet.append(row)

    # Create a new sheet '1_Working'
    working_sheet = wb.create_sheet('1_Working')
    
    # Get the "1. Course Topic" sheet
    topic_sheet = wb['1. Course Topic']
    
    # Copy values from columns A, B, and C in '1. Course Topic' to columns W, X, and Y in 'Mappings'
    for row in range(1, max_row + 1):
        mappings_sheet.cell(row=row, column=23, value=topic_sheet.cell(row=row, column=1).value)
        mappings_sheet.cell(row=row, column=24, value=topic_sheet.cell(row=row, column=2).value)
        mappings_sheet.cell(row=row, column=25, value=topic_sheet.cell(row=row, column=3).value)

    # Copy columns B and C from '1. Course Topic' to '1_Working'
    for row in range(1, max_row + 1):
        working_sheet[f'A{row}'] = topic_sheet[f'B{row}'].value
        working_sheet[f'B{row}'] = topic_sheet[f'C{row}'].value
    
    # Insert the formula =trim(B1) in column C1 onwards while cells of column B is not empty
    row = 1
    while working_sheet[f'B{row}'].value:
        working_sheet[f'C{row}'] = f'=TRIM(B{row})'
        row += 1
    
    # Create a new sheet '1_CSV'
    csv_sheet_1 = wb.create_sheet('1_CSV')

    # Create a new sheet '2_CSV'
    csv_sheet_2 = wb.create_sheet('2_CSV')
    
    # Get the "2. Course Paper" sheet
    course_paper_sheet = wb['2. Course Paper']
    
    # Copy values from columns B to D from '2. Course Paper' to '2_CSV'
    for row in range(1, max_row + 1):
        csv_sheet_2.cell(row=row, column=1, value=course_paper_sheet[f'B{row}'].value)
        csv_sheet_2.cell(row=row, column=2, value=course_paper_sheet[f'C{row}'].value)
        csv_sheet_2.cell(row=row, column=3, value=course_paper_sheet[f'D{row}'].value)

    # Create a new sheet '3_Working'
    working_sheet_3 = wb.create_sheet('3_Working')
    
    # Get the "3. Course Template" sheet
    template_sheet = wb['3. Course Template']
    
    # Copy data from '3. Course Template' to '3_Working'
    for row in range(1, max_row + 1):
        for col in range(1, template_sheet.max_column + 1):
            working_sheet_3.cell(row=row, column=col, value=template_sheet.cell(row=row, column=col).value)
    
    # Insert Course ID, Level ID, and Cohort headers
    working_sheet_3['G1'] = 'Course ID'
    working_sheet_3['H1'] = 'Level ID'
    working_sheet_3['I1'] = 'Cohort'
    
    # Insert formulas in G2, H2, and I2
    working_sheet_3[f'G2'] = f'=VLOOKUP(B2, Mappings!A:D, 4, FALSE)'
    working_sheet_3[f'H2'] = f'=VLOOKUP(C2, Mappings!G:I, 3, FALSE)'
    working_sheet_3[f'I2'] = f'=VLOOKUP(MID(A2, FIND("-", A2)+1, 4), \'Cohort Mapping\'!$A:$B, 2, FALSE)'

    # Create a new sheet '3_CSV'
    csv_sheet_3 = wb.create_sheet('3_CSV')

    # Save the edited Excel file
    wb.save(filename)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if file_path:
        edit_excel_file(file_path)
        print('File successfully edited and saved.')

if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    select_file()
