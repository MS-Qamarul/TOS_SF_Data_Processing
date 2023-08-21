import openpyxl
import tkinter as tk
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import pandas as pd

def edit_excel_file(filename):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)
    mappings_sheet = workbook['Mappings']
    course_question_sheet = workbook['4. Course Question']

    # Step 1: Insert "=AB1", "=AB2", ... into cells in column AF while AB is not empty
    row = 1
    while mappings_sheet[f'AB{row}'].value:
        mappings_sheet[f'AF{row}'] = f'=AB{row}'
        row += 1

    # Step 2: Insert "Remapped Course Topic Name" into cell Z1
    mappings_sheet['Z1'] = 'Remapped Course Topic Name'

    # Step 3: Insert VLOOKUP formula into cells in column Z while X is not empty
    row = 2
    while mappings_sheet[f'X{row}'].value:
        mappings_sheet[f'Z{row}'] = f'=VLOOKUP(TRIM(Y{row}), AD:AF, 3, FALSE)'
        row += 1

    # Step 4: Create a new sheet called "4_Working"
    working_sheet = workbook.create_sheet("4_Working")

    # Step 5: Copy values from columns A to K from "4. Course Question" sheet to "4_Working" sheet
    for row_index, row in enumerate(course_question_sheet.iter_rows(min_row=1, max_row=course_question_sheet.max_row, min_col=1, max_col=11), start=1):
        for col_index, cell in enumerate(row, start=1):
            dest_cell = working_sheet.cell(row=row_index, column=col_index, value=cell.value)

    # Step 6: Insert headers into cells L1, M1, N1, O1, and P1 in "4_Working" sheet
    working_sheet['L1'] = 'Remapped Course Topic Name'
    working_sheet['M1'] = 'Course Template ID'
    working_sheet['N1'] = 'Course Topic ID'
    working_sheet['O1'] = 'Course Paper ID'
    working_sheet['P1'] = 'Seq Number'

    # Step 7: Insert formulas into columns L, M, N, and O in "4_Working" sheet
    row = 2
    while working_sheet[f'A{row}'].value:
        working_sheet[f'L{row}'] = f'=VLOOKUP(D{row}, Mappings!W:Z, 4, FALSE)'
        working_sheet[f'M{row}'] = f'=VLOOKUP(A{row}, Mappings!R:U, 4, FALSE)'
        working_sheet[f'N{row}'] = f'=VLOOKUP(L{row}, Mappings!AB:AE, 4, FALSE)'
        working_sheet[f'O{row}'] = f'=VLOOKUP(J{row}, Mappings!L:P, 5, FALSE)'
        row += 1

    # Step 8: Create a new sheet called "4_CSV"
    csv_sheet = workbook.create_sheet("4_CSV")

    # Save the modified Excel file
    workbook.save(filename)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if file_path:
        edit_excel_file(file_path)
        print('File successfully edited and saved.')

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    select_file()
